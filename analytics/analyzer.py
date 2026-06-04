from __future__ import annotations
import csv
import os
import io
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from config import DATA_DIR, PREFIX_MAP_FILE

_THRESHOLDS_FILE = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'cbse', 'thresholds.json')

def _load_thresholds() -> dict:
    with open(_THRESHOLDS_FILE, 'r') as f:
        return json.load(f)

_THRESHOLDS = _load_thresholds()


def load_results_csv(school_no: str) -> tuple[list[dict] | None, dict[str, str] | None]:
    students, subjects_map = {}, {}
    csv_file = os.path.join(DATA_DIR, f"{school_no}_results.csv")
    if not os.path.exists(csv_file):
        return None, None

    prefix_map = {}
    if os.path.exists(PREFIX_MAP_FILE):
        try:
            with open(PREFIX_MAP_FILE, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    prefix_map[int(row['Roll'].strip())] = row['AdmitCardID'].strip().upper()
        except Exception:
            pass

    try:
        with open(csv_file, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                roll = int(row['Roll'].strip())
                name, mother, father, school = (row.get(k, '').strip().upper() for k in ['Name', 'MotherName', 'FatherName', 'SchoolName'])
                
                sub_code = row['SubjectCode'].strip().upper()
                sub_name = row['SubjectName'].strip().upper()
                t_str, p_str = row['Theory'].strip(), row['Practical'].strip()
                t_marks = int(t_str) if t_str.isdigit() else 0
                p_marks = int(p_str) if p_str.isdigit() else 0
                result_status = row['Result'].strip().upper()

                if sub_code not in subjects_map:
                    subjects_map[sub_code] = sub_name

                if roll not in students:
                    students[roll] = {
                        'roll': roll, 'name': name, 'mother': mother, 'father': father,
                        'school': school, 'admid': prefix_map.get(roll, "N/A"),
                        'marks': {}, 'result': result_status
                    }
                students[roll]['marks'][sub_code] = {'t': t_marks, 'p': p_marks}
    except Exception:
        return None, None

    return list(students.values()), subjects_map


def process_student_data(students: list[dict], subjects_map: dict[str, str]) -> list[dict]:
    scored_students = []
    prac_max_map = _THRESHOLDS['practical_max']
    skip_subs = _THRESHOLDS['skip_subjects']
    
    for s in students:
        total_score, count, fail_count_sub = 0, 0, 0
        
        for sub_code in s['marks']:
            if sub_code in skip_subs:
                continue
                
            t_mark = s['marks'][sub_code]['t']
            p_mark = s['marks'][sub_code]['p']
            sub_total = t_mark + p_mark
            total_score += sub_total
            count += 1
            
            max_p = prac_max_map.get(sub_code, prac_max_map['default'])
            max_t = 100 - max_p
            
            # CBSE pass criteria: min 33% overall, plus subject-specific theory/practical thresholds
            pass_t = 26 if max_t == 80 else (23 if max_t == 70 else (10 if max_t == 30 else (20 if max_t == 60 else 33)))
            pass_p = 7 if max_p == 20 else (10 if max_p == 30 else (23 if max_p == 70 else (13 if max_p == 40 else 0)))
            
            subject_passed = True
            if sub_total < 33: subject_passed = False
            if max_t > 0 and t_mark < pass_t: subject_passed = False
            if max_p > 0 and p_mark < pass_p: subject_passed = False
            if not subject_passed: fail_count_sub += 1

        if fail_count_sub == 0: s['result'] = 'PASS'
        elif fail_count_sub <= 2: s['result'] = 'COMP'
        else: s['result'] = 'FAIL'

        percentage = (total_score / count) if count > 0 else 0
        s['percentage'] = percentage
        s['totalScore'] = total_score
        
        if percentage >= 90: s['category'] = "Excellent (>=90%)"
        elif percentage >= 75: s['category'] = "Good (75-89%)"
        elif percentage >= 50: s['category'] = "Average (50-74%)"
        else: s['category'] = "Needs Attention (<50%)"

        scored_students.append(s)

    scored_students.sort(key=lambda x: x['percentage'], reverse=True)
    return scored_students


def get_dashboard_data(school_no: str) -> dict:
    students, subjects_map = load_results_csv(school_no)
    if not students:
        return {"error": "No data found."}
    
    processed_students = process_student_data(students, subjects_map)
    return {
        "students": processed_students,
        "subjects": subjects_map
    }


def delete_database(school_no: str) -> bool:
    csv_file = os.path.join(DATA_DIR, f"{school_no}_results.csv")
    if os.path.exists(csv_file):
        try:
            os.remove(csv_file)
            return True
        except Exception:
            return False
    return True


def delete_record(roll_to_delete: str, school_no: str) -> bool:
    csv_file = os.path.join(DATA_DIR, f"{school_no}_results.csv")
    if not os.path.exists(csv_file):
        return False
    
    try:
        rows_to_keep = []
        deleted = False
        with open(csv_file, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            headers = next(reader, None)
            if headers:
                rows_to_keep.append(headers)
            for row in reader:
                if len(row) > 0 and str(row[0]).strip() == str(roll_to_delete):
                    deleted = True
                else:
                    rows_to_keep.append(row)
                    
        if deleted:
            with open(csv_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerows(rows_to_keep)
        return deleted
    except Exception:
        return False


def generate_excel_bytes(school_no: str) -> io.BytesIO | None:
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.series import DataPoint
    from openpyxl.utils import get_column_letter

    students, subjects_map = load_results_csv(school_no)
    if not students:
        return None

    processed = process_student_data(students, subjects_map)

    INDIGO      = "4F46E5"
    DARK_BG     = "1E1B4B"
    LIGHT_BG    = "EEF2FF"
    WHITE       = "FFFFFF"
    PASS_GREEN  = "16A34A"
    FAIL_RED    = "DC2626"
    COMP_ORANGE = "D97706"
    BORDER_CLR  = "C7D2FE"
    ALT_ROW     = "F5F3FF"

    title_font   = Font(color=WHITE, bold=True, size=18)
    title_fill   = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
    section_font = Font(bold=True, size=13, color=DARK_BG)
    header_font  = Font(color=WHITE, bold=True, size=11)
    header_fill  = PatternFill(start_color=INDIGO, end_color=INDIGO, fill_type="solid")
    stat_lbl     = Font(bold=True, size=10, color="6B7280")
    stat_val     = Font(bold=True, size=16, color=INDIGO)
    pass_font    = Font(color=PASS_GREEN, bold=True)
    fail_font    = Font(color=FAIL_RED, bold=True)
    comp_font    = Font(color=COMP_ORANGE, bold=True)
    center       = Alignment(horizontal="center", vertical="center")
    left_al      = Alignment(horizontal="left", vertical="center")
    wrap_center  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    light_fill   = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
    alt_fill     = PatternFill(start_color=ALT_ROW, end_color=ALT_ROW, fill_type="solid")
    thin_border  = Border(
        left=Side(style='thin', color=BORDER_CLR), right=Side(style='thin', color=BORDER_CLR),
        top=Side(style='thin', color=BORDER_CLR), bottom=Side(style='thin', color=BORDER_CLR))

    def set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def write_header_row(ws, row, headers):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = wrap_center
            c.border = thin_border

    def style_result(cell, val):
        if val == "PASS": cell.font = pass_font
        elif val == "FAIL": cell.font = fail_font
        else: cell.font = comp_font

    def auto_width(ws, col_count):
        for col in range(1, col_count + 1):
            letter = get_column_letter(col)
            mx = 0
            for row in ws[letter]:
                if row.value:
                    mx = max(mx, len(str(row.value)))
            ws.column_dimensions[letter].width = min(mx + 3, 40)

    scoring_codes = sorted([c for c in subjects_map if c not in ['500', '502', '503']])
    total   = len(processed)
    pass_c  = sum(1 for s in processed if s['result'] == 'PASS')
    comp_c  = sum(1 for s in processed if s['result'] == 'COMP')
    fail_c  = sum(1 for s in processed if s['result'] == 'FAIL')
    avg_pct = sum(s['percentage'] for s in processed) / total if total else 0
    topper  = processed[0]['name'] if processed else "N/A"

    cats = {"Excellent (>=90%)": 0, "Good (75-89%)": 0, "Average (50-74%)": 0, "Needs Attention (<50%)": 0}
    for s in processed:
        if s['category'] in cats:
            cats[s['category']] += 1

    stream_map = {}
    for s in processed:
        codes = set(s['marks'].keys()) - {'500', '502', '503', '301'}
        if codes & {'042', '043'}:
            stream_map.setdefault("Science", []).append(s)
        elif codes & {'055', '054'}:
            stream_map.setdefault("Commerce", []).append(s)
        else:
            stream_map.setdefault("Humanities", []).append(s)

    wb = openpyxl.Workbook()

    # Sheet 1: Dashboard
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = DARK_BG

    set_col_widths(ws, [18, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14])

    ws.merge_cells('A1:L2')
    tc = ws['A1']
    tc.value = f"School {school_no} — Results Dashboard"
    tc.font = title_font
    tc.fill = title_fill
    tc.alignment = center
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 28
    for col in range(1, 13):
        ws.cell(row=1, column=col).fill = title_fill
        ws.cell(row=2, column=col).fill = title_fill

    stat_items = [
        ("TOTAL STUDENTS", total),
        ("PASSED", pass_c),
        ("COMPARTMENT", comp_c),
        ("FAILED", fail_c),
        ("AVERAGE %", f"{avg_pct:.1f}%"),
        ("TOPPER", topper),
    ]
    for i, (label, val) in enumerate(stat_items):
        col_start = i * 2 + 1
        c1 = ws.cell(row=4, column=col_start, value=label)
        c1.font = stat_lbl
        c1.alignment = center
        c1.fill = light_fill
        c1.border = thin_border
        ws.merge_cells(start_row=4, start_column=col_start, end_row=4, end_column=col_start + 1)
        ws.cell(row=4, column=col_start + 1).fill = light_fill
        ws.cell(row=4, column=col_start + 1).border = thin_border
        c2 = ws.cell(row=5, column=col_start, value=val)
        c2.font = stat_val
        c2.alignment = center
        c2.fill = light_fill
        c2.border = thin_border
        ws.merge_cells(start_row=5, start_column=col_start, end_row=5, end_column=col_start + 1)
        ws.cell(row=5, column=col_start + 1).fill = light_fill
        ws.cell(row=5, column=col_start + 1).border = thin_border
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 30

    ws.cell(row=7, column=1, value="Result Breakdown").font = section_font
    ws.cell(row=7, column=7, value="Performance Distribution").font = section_font

    for i, h in enumerate(["Category", "Count"], 1):
        c = ws.cell(row=8, column=i, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = center; c.border = thin_border
    pie_items = [("Pass", pass_c), ("Compartment", comp_c), ("Fail", fail_c)]
    for r, (cat, cnt) in enumerate(pie_items, 9):
        ws.cell(row=r, column=1, value=cat).border = thin_border
        ws.cell(row=r, column=2, value=cnt).border = thin_border
        ws.cell(row=r, column=1).alignment = left_al
        ws.cell(row=r, column=2).alignment = center

    for i, h in enumerate(["Category", "Students"], 7):
        c = ws.cell(row=8, column=i, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = center; c.border = thin_border
    for r, (cat, cnt) in enumerate(cats.items(), 9):
        ws.cell(row=r, column=7, value=cat).border = thin_border
        ws.cell(row=r, column=8, value=cnt).border = thin_border
        ws.cell(row=r, column=7).alignment = left_al
        ws.cell(row=r, column=8).alignment = center

    pie = PieChart()
    pie.title = "Pass / Fail / Compartment"
    pie.style = 10
    pie.width = 14
    pie.height = 10
    labels = Reference(ws, min_col=1, min_row=9, max_row=11)
    data   = Reference(ws, min_col=2, min_row=8, max_row=11)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = True
    green_pt = DataPoint(idx=0); green_pt.graphicalProperties.solidFill = PASS_GREEN
    orange_pt = DataPoint(idx=1); orange_pt.graphicalProperties.solidFill = COMP_ORANGE
    red_pt = DataPoint(idx=2); red_pt.graphicalProperties.solidFill = FAIL_RED
    pie.series[0].data_points = [green_pt, orange_pt, red_pt]
    ws.add_chart(pie, "A13")

    bar = BarChart()
    bar.type = "col"
    bar.title = "Grade Distribution"
    bar.style = 10
    bar.width = 14
    bar.height = 10
    bar.y_axis.title = "Students"
    bar_data = Reference(ws, min_col=8, min_row=8, max_row=12)
    bar_cats = Reference(ws, min_col=7, min_row=9, max_row=12)
    bar.add_data(bar_data, titles_from_data=True)
    bar.set_categories(bar_cats)
    bar.series[0].graphicalProperties.solidFill = INDIGO
    ws.add_chart(bar, "G13")

    tbl_row = 29
    ws.cell(row=tbl_row, column=1, value="Subject-wise Performance").font = section_font
    tbl_row += 1
    subj_headers = ["Subject", "Avg Marks", "Highest", "Lowest", "Students"]
    write_header_row(ws, tbl_row, subj_headers)
    tbl_row += 1
    for code in scoring_codes:
        name = subjects_map.get(code, code)
        totals = [s['marks'][code]['t'] + s['marks'][code]['p'] for s in processed if code in s['marks']]
        if not totals:
            continue
        row_data = [name, round(sum(totals)/len(totals), 1), max(totals), min(totals), len(totals)]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=tbl_row, column=col, value=val)
            c.border = thin_border
            c.alignment = center if col > 1 else left_al
        tbl_row += 1

    tbl_row += 1
    ws.cell(row=tbl_row, column=1, value="Top 10 Students").font = section_font
    tbl_row += 1
    top_headers = ["Rank", "Roll No", "Name", "Total Score", "Percentage", "Result"]
    write_header_row(ws, tbl_row, top_headers)
    tbl_row += 1
    for rank, s in enumerate(processed[:10], 1):
        row_data = [rank, s['roll'], s['name'], s['totalScore'], f"{s['percentage']:.1f}%", s['result']]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=tbl_row, column=col, value=val)
            c.border = thin_border
            c.alignment = center if col != 3 else left_al
            if col == 6:
                style_result(c, val)
        tbl_row += 1

    # Sheet 2: Full Rankings
    ws2 = wb.create_sheet("Full Rankings")
    ws2.sheet_properties.tabColor = INDIGO

    hdrs2 = ["Rank", "Roll No", "Admit ID", "Name", "Father Name", "Mother Name", "Total", "%", "Result"]
    for code in scoring_codes:
        hdrs2.append(subjects_map.get(code, code))

    write_header_row(ws2, 1, hdrs2)
    ws2.row_dimensions[1].height = 30
    ws2.freeze_panes = 'A2'

    for ri, s in enumerate(processed, 2):
        row = [ri - 1, s['roll'], s['admid'], s['name'], s['father'], s['mother'],
               s['totalScore'], round(s['percentage'], 2), s['result']]
        for code in scoring_codes:
            if code in s['marks']:
                row.append(s['marks'][code]['t'] + s['marks'][code]['p'])
            else:
                row.append("-")

        is_alt = (ri % 2 == 0)
        for col, val in enumerate(row, 1):
            c = ws2.cell(row=ri, column=col, value=val)
            c.border = thin_border
            if is_alt:
                c.fill = alt_fill
            c.alignment = center if col in [1, 2, 7, 8, 9] else left_al
            if col == 9:
                style_result(c, val)

    auto_width(ws2, len(hdrs2))

    stream_colors = {"Science": "0EA5E9", "Commerce": "8B5CF6", "Humanities": "F59E0B"}
    for stream_name, stream_students in stream_map.items():
        if not stream_students:
            continue

        ws_s = wb.create_sheet(stream_name)
        ws_s.sheet_properties.tabColor = stream_colors.get(stream_name, INDIGO)
        clr = stream_colors.get(stream_name, INDIGO)

        stream_sub_codes = sorted({c for s in stream_students for c in s['marks'] if c not in ['500', '502', '503']})

        ws_s.merge_cells('A1:J1')
        tc = ws_s['A1']
        tc.value = f"{stream_name} Stream — {len(stream_students)} Students"
        tc.font = Font(color=WHITE, bold=True, size=14)
        tc.fill = PatternFill(start_color=clr, end_color=clr, fill_type="solid")
        tc.alignment = center
        ws_s.row_dimensions[1].height = 36
        for col in range(1, 11):
            ws_s.cell(row=1, column=col).fill = PatternFill(start_color=clr, end_color=clr, fill_type="solid")

        s_pass = sum(1 for s in stream_students if s['result'] == 'PASS')
        s_avg  = sum(s['percentage'] for s in stream_students) / len(stream_students)
        stats_items = [("Students", len(stream_students)), ("Passed", s_pass),
                       ("Pass Rate", f"{s_pass/len(stream_students)*100:.0f}%"), ("Avg %", f"{s_avg:.1f}%")]
        for i, (lbl, val) in enumerate(stats_items):
            col_s = i * 2 + 1
            c1 = ws_s.cell(row=3, column=col_s, value=lbl)
            c1.font = stat_lbl; c1.fill = light_fill; c1.border = thin_border; c1.alignment = center
            c2 = ws_s.cell(row=3, column=col_s + 1, value=val)
            c2.font = stat_val; c2.fill = light_fill; c2.border = thin_border; c2.alignment = center

        hdrs_s = ["Rank", "Roll No", "Name", "Total", "%", "Result"]
        for code in stream_sub_codes:
            hdrs_s.append(subjects_map.get(code, code))

        data_row = 5
        write_header_row(ws_s, data_row, hdrs_s)
        ws_s.row_dimensions[data_row].height = 28
        ws_s.freeze_panes = f'A{data_row + 1}'

        for ri, s in enumerate(stream_students, data_row + 1):
            row = [ri - data_row, s['roll'], s['name'], s['totalScore'],
                   round(s['percentage'], 2), s['result']]
            for code in stream_sub_codes:
                if code in s['marks']:
                    row.append(s['marks'][code]['t'] + s['marks'][code]['p'])
                else:
                    row.append("-")

            is_alt = ((ri - data_row) % 2 == 0)
            for col, val in enumerate(row, 1):
                c = ws_s.cell(row=ri, column=col, value=val)
                c.border = thin_border
                if is_alt:
                    c.fill = alt_fill
                c.alignment = center if col in [1, 2, 4, 5, 6] else left_al
                if col == 6:
                    style_result(c, val)

        chart_start = data_row + len(stream_students) + 2
        ws_s.cell(row=chart_start, column=1, value="Subject Averages").font = section_font
        chart_start += 1
        ws_s.cell(row=chart_start, column=1, value="Subject").font = header_font
        ws_s.cell(row=chart_start, column=1).fill = header_fill
        ws_s.cell(row=chart_start, column=1).border = thin_border
        ws_s.cell(row=chart_start, column=2, value="Average").font = header_font
        ws_s.cell(row=chart_start, column=2).fill = header_fill
        ws_s.cell(row=chart_start, column=2).border = thin_border

        cr = chart_start + 1
        for code in stream_sub_codes:
            totals = [s['marks'][code]['t'] + s['marks'][code]['p'] for s in stream_students if code in s['marks']]
            if totals:
                ws_s.cell(row=cr, column=1, value=subjects_map.get(code, code)).border = thin_border
                ws_s.cell(row=cr, column=2, value=round(sum(totals)/len(totals), 1)).border = thin_border
                cr += 1

        if cr > chart_start + 1:
            sb = BarChart()
            sb.type = "col"
            sb.title = f"{stream_name} — Subject Averages"
            sb.style = 10
            sb.width = 16
            sb.height = 10
            sb.y_axis.title = "Average Marks"
            sb_data = Reference(ws_s, min_col=2, min_row=chart_start, max_row=cr - 1)
            sb_cats = Reference(ws_s, min_col=1, min_row=chart_start + 1, max_row=cr - 1)
            sb.add_data(sb_data, titles_from_data=True)
            sb.set_categories(sb_cats)
            sb.series[0].graphicalProperties.solidFill = clr
            ws_s.add_chart(sb, f"D{chart_start}")

        auto_width(ws_s, len(hdrs_s))

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out
